from io import BytesIO
from typing import List
import openpyxl
from fastapi import Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from core.database import get_db


class ExportRequest(BaseModel):
    candidate_ids: List[str]


async def export_shortlist(
    request: ExportRequest,
    db: AsyncSession = Depends(get_db)
):
    # Fetch candidates
    sql = text("""
        SELECT 
            candidate_id,
            first_name || ' ' || last_name AS full_name,
            seniority_level,
            location_city,
            availability_status,
            years_of_experience,
            salary_expectation_min,
            salary_expectation_max,
            skills_primary,
            business_domains,
            job_titles_canonical,
            holistic_summary_text
        FROM candidates
        WHERE candidate_id = ANY(:ids)
    """)
    result = await db.execute(
        sql, {"ids": request.candidate_ids}
    )
    rows = result.fetchall()

    # Seniority labels
    seniority_map = {
        0: "Fresher", 1: "Junior", 2: "Mid",
        3: "Senior", 4: "Lead", 5: "Expert"
    }

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # Headers
    headers = [
        "Full Name", "Seniority", "Location",
        "Availability", "Years Experience",
        "Salary Min (VND)", "Salary Max (VND)",
        "Primary Skills", "Business Domains",
        "Job Titles", "Summary", "CV Link"
    ]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="C0392B",
            end_color="C0392B",
            fill_type="solid"
        )

    # Data rows
    for row in rows:
        # Dummy CV link — replace with real storage URL if available
        cv_link = f"https://example.com/cv/{row.candidate_id}"
        ws.append([
            row.full_name,
            seniority_map.get(row.seniority_level, "Unknown"),
            row.location_city,
            row.availability_status,
            row.years_of_experience,
            row.salary_expectation_min,
            row.salary_expectation_max,
            ", ".join(row.skills_primary or []),
            ", ".join(row.business_domains or []),
            ", ".join(row.job_titles_canonical or []),
            row.holistic_summary_text or "",
            cv_link
        ])

    # Auto column width
    for col in ws.columns:
        max_len = max(
            len(str(cell.value or "")) for cell in col
        )
        ws.column_dimensions[
            col[0].column_letter
        ].width = min(max_len + 2, 50)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 
            "attachment; filename=candidates_export.xlsx"
        }
    )
